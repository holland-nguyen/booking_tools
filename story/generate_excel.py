#!/usr/bin/env python3
"""
Generate two Excel files for Golf Booking Project:
1. Questions / Clarification List (for client)
2. Quotation & Timeline (for client)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
import os

OUTPUT_DIR = "/Users/hacnguyen/Documents/booking_tools/story"

# ─── Color Palette ────────────────────────────────────────────────────────────
GREEN_DARK   = "1A472A"
GREEN_MID    = "2D6A4F"
GREEN_LIGHT  = "52B788"
GREEN_PALE   = "D8F3DC"
GOLD         = "B5860D"
GOLD_LIGHT   = "FFF3CD"
GRAY_DARK    = "495057"
GRAY_LIGHT   = "F8F9FA"
WHITE        = "FFFFFF"
RED_SOFT     = "C0392B"
ORANGE_SOFT  = "E67E22"
BLUE_SOFT    = "2471A3"

def thin_border():
    thin = Side(style='thin', color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ══════════════════════════════════════════════════════════════════════════════
#  FILE 1: QUESTIONS / CLARIFICATION LIST
# ══════════════════════════════════════════════════════════════════════════════
def create_questions_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Câu hỏi làm rõ"
    ws.sheet_view.showGridLines = False

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "GOLF TEE TIME BOOKING SYSTEM"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    ws["A1"].fill = header_fill(GREEN_DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = "Danh sách câu hỏi làm rõ yêu cầu — Requirements Clarification"
    ws["A2"].font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws["A2"].fill = header_fill(GREEN_MID)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    ws.merge_cells("A3:G3")
    ws["A3"] = "Phiên bản: 1.0  |  Ngày lập: 15/04/2026  |  Đơn vị phát triển: [Tên công ty]  |  Khách hàng: [Tên sân golf]"
    ws["A3"].font = Font(italic=True, size=9, color=GRAY_DARK, name="Calibri")
    ws["A3"].fill = header_fill(GREEN_PALE)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    ws.row_dimensions[4].height = 8  # spacer

    # ── Column headers ────────────────────────────────────────────────────────
    headers = ["STT", "Chủ đề", "Câu hỏi", "Lý do cần làm rõ", "Mức ưu tiên", "Phản hồi của khách", "Ghi chú"]
    col_widths = [6, 22, 52, 38, 12, 30, 20]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
        cell.fill = header_fill(GREEN_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        set_col_width(ws, get_column_letter(i), w)

    ws.row_dimensions[5].height = 30

    # ── Priority styling ──────────────────────────────────────────────────────
    priority_style = {
        "Cao": (RED_SOFT, WHITE),
        "Trung bình": (ORANGE_SOFT, WHITE),
        "Thấp": (BLUE_SOFT, WHITE),
    }

    # ── Data ──────────────────────────────────────────────────────────────────
    questions = [
        # STT, Group, Question, Reason, Priority
        # === GROUP A ===
        ("A", "CẤU HÌNH SÂN & LỊCH ĐẶT", "", "", ""),
        ("A1", "Cấu hình sân", "Khoảng cách (interval) giữa các tee time là bao nhiêu phút? Con số này có khác nhau giữa 1st Tee và 10th Tee không?", "Ảnh hưởng trực tiếp đến cấu trúc dữ liệu và giao diện bảng tee sheet.", "Cao"),
        ("A2", "Cấu hình sân", "Giờ hoạt động của sân mỗi ngày trong tuần là từ mấy giờ đến mấy giờ? Có sự khác biệt giữa các ngày trong tuần và cuối tuần không?", "Xác định khung giờ hợp lệ để hiển thị trên hệ thống đặt lịch.", "Cao"),
        ("A3", "Cấu hình sân", "Khi đặt trực tuyến, khách hàng có thể tự chọn 1st Tee hoặc 10th Tee không? Hay hệ thống sẽ tự động phân bổ?", "Quyết định luồng đặt lịch và giao diện widget.", "Cao"),
        ("A4", "Giá dịch vụ", "Giá tee time được tính theo cấu trúc nào? Cụ thể:\n  • Ngày thường và cuối tuần/lễ có giá khác nhau không?\n  • Phân biệt giờ cao điểm và thấp điểm không?\n  • Hội viên và khách vãng lai có biểu giá khác nhau không?\n  • Quốc tịch (trong nước / nước ngoài) có ảnh hưởng đến giá không?", "Thiết kế module tính giá và hiển thị tổng tiền trước khi thanh toán.", "Cao"),
        # === GROUP B ===
        ("B", "THANH TOÁN & CHÍNH SÁCH", "", "", ""),
        ("B1", "Thanh toán", "Khi khách xác nhận đặt lịch và thanh toán, hệ thống sẽ xử lý theo phương thức nào?\n  • Phương án 1 (Authorize): Tạm giữ số tiền trên thẻ, chờ nhân viên xác nhận rồi mới thu.\n  • Phương án 2 (Capture ngay): Thu tiền ngay lập tức khi đặt thành công.", "Quyết định này ảnh hưởng trực tiếp đến quy trình thanh toán và hoàn tiền.", "Cao"),
        ("B2", "Hoàn tiền", "Chính sách hoàn tiền khi khách hủy đặt lịch như thế nào?\n  • Hoàn 100% nếu hủy trước X giờ?\n  • Hoàn một phần nếu hủy muộn?\n  • Không hoàn tiền?", "Cần xây dựng quy trình hoàn tiền tự động qua cổng thanh toán.", "Cao"),
        ("B3", "Thanh toán", "Tiền tệ sử dụng trong hệ thống là VND, USD hay hỗ trợ cả hai?", "Cấu hình cổng thanh toán Airwallex.", "Cao"),
        ("B4", "Thanh toán", "Nhân viên khi tự tạo booking cho khách (walk-in, qua điện thoại) có cần xử lý thanh toán trực tuyến không? Hay thanh toán trực tiếp tại quầy?", "Phân biệt luồng booking online và booking thủ công.", "Trung bình"),
        # === GROUP C ===
        ("C", "THÔNG TIN KHÁCH HÀNG & TEE SHEET", "", "", ""),
        ("C1", "Dữ liệu khách", "Trong bảng tee sheet cũ, mỗi slot hiển thị tên từng người chơi (Player 1, 2, 3, 4). Hiện tại form đặt trực tuyến chỉ thu email và số điện thoại. Vậy tên người chơi được lấy từ đâu?\n  • Nhân viên nhập thủ công sau khi gọi điện xác nhận?\n  • Cần bổ sung trường tên người chơi vào form thanh toán?", "Ảnh hưởng đến thiết kế form thu thập dữ liệu và hiển thị tee sheet.", "Cao"),
        ("C2", "Xác nhận", "Sau khi khách đặt lịch trực tuyến, nhân viên có bao nhiêu thời gian để gọi điện xác nhận? Nếu quá thời gian quy định mà chưa xác nhận thì booking sẽ được xử lý thế nào (tự hủy hay tự xác nhận)?", "Xây dựng logic tự động hóa vòng đời booking.", "Trung bình"),
        ("C3", "Khách hàng", "Khách có thể vào danh sách chờ (waiting list) khi slot đã đủ 4 người không? Nếu có: khi có người hủy thì nhân viên thông báo thủ công hay hệ thống tự động thông báo?", "Thiết kế tính năng waiting list.", "Trung bình"),
        # === GROUP D ===
        ("D", "TÍNH NĂNG ADMIN — CẦN LÀM RÕ", "", "", ""),
        ("D1", "Tính năng", "\"Register\" trong hệ thống cũ thực hiện chức năng gì? Đơn vị hiểu có thể là:\n  (A) Đăng ký thêm dịch vụ kèm theo (caddy, xe điện, thuê gậy...)\n  (B) Đăng ký tư cách hội viên cho khách\n  (C) Chức năng khác — đề nghị mô tả cụ thể.", "Xác định phạm vi tính năng.", "Trung bình"),
        ("D2", "Tính năng", "\"Extra Item\" bao gồm những hạng mục nào? Ví dụ: phí caddy, xe điện, thuê bộ gậy, locker... Danh sách đầy đủ các mục hiện tại?", "Thiết kế module quản lý dịch vụ đi kèm.", "Trung bình"),
        ("D3", "Tính năng", "\"Allocate Info\" thực hiện việc gì? Đơn vị hiểu có thể là việc gán thông tin bổ sung vào booking sau khi xác nhận (caddy số mấy, xe điện số mấy...). Quý khách vui lòng mô tả thêm?", "Xác định phạm vi tính năng.", "Trung bình"),
        ("D4", "Tính năng", "\"Squeeze\" và \"Unsqueeze\" hoạt động như thế nào trong quy trình hiện tại? Đây có phải là việc thêm một slot vào giữa các tee time đang đặt (rút ngắn khoảng cách) không?", "Xác định phạm vi tính năng.", "Trung bình"),
        ("D5", "Tính năng", "\"Full Block\" khác gì với \"Block Tee Time\" thông thường? Có phải là chặn toàn bộ một tee (1st hoặc 10th) trong một khoảng thời gian dài không?", "Xác định phạm vi tính năng.", "Trung bình"),
        ("D6", "Tính năng", "\"Time Sheet\" trong hệ thống cũ là màn hình hoặc báo cáo gì? Khác với bảng Tee Sheet ở điểm nào?", "Xác định phạm vi tính năng.", "Trung bình"),
        ("D7", "Tính năng", "\"Confirm Letter\" được gửi đến khách qua kênh nào (email, PDF tải về, in trực tiếp)? Template hiện tại có thể cung cấp để tham khảo không?", "Thiết kế tính năng gửi xác nhận.", "Thấp"),
        ("D8", "Tính năng", "\"Group Booking\" có áp dụng mức giá đặc biệt không? Quy trình phê duyệt như thế nào?", "Thiết kế luồng nhóm booking.", "Trung bình"),
        # === GROUP E ===
        ("E", "PHÂN QUYỀN & VẬN HÀNH", "", "", ""),
        ("E1", "Phân quyền", "Nhân viên Sale có quyền tự xác nhận (confirm) booking của khách đặt online không? Hay quy trình phải qua Admin phê duyệt?", "Xây dựng ma trận phân quyền chính xác.", "Trung bình"),
        ("E2", "Phân quyền", "Nhân viên Support có những quyền hạn cụ thể nào? Chỉ xem thông tin hay có thêm quyền thao tác?", "Xây dựng ma trận phân quyền chính xác.", "Trung bình"),
        ("E3", "Phân quyền", "Có yêu cầu ghi lại lịch sử thao tác (audit log) — ai làm gì, lúc nào — không?", "Xác định phạm vi tính năng bảo mật.", "Thấp"),
        # === GROUP F ===
        ("F", "KỸ THUẬT & TRIỂN KHAI", "", "", ""),
        ("F1", "Kỹ thuật", "Ngôn ngữ hiển thị của hệ thống: tiếng Việt, tiếng Anh hay hỗ trợ đa ngôn ngữ?", "Phạm vi giao diện.", "Thấp"),
        ("F2", "Kỹ thuật", "Sân có đang sử dụng phần mềm quản lý nào khác không (POS, hệ thống hội viên, tính điểm handicap...)? Có yêu cầu tích hợp không?", "Xác định phạm vi tích hợp.", "Thấp"),
        ("F3", "Kỹ thuật", "Dữ liệu booking lịch sử (từ hệ thống cũ) có cần chuyển đổi sang hệ thống mới không?", "Lập kế hoạch di chuyển dữ liệu.", "Thấp"),
        ("F4", "Tài nguyên", "Đơn vị cần nhận: logo sân golf, ảnh sân golf (cho slider), màu sắc thương hiệu (brand color). Quý khách có thể cung cấp sớm để thiết kế giao diện không?", "Bắt đầu thiết kế UI.", "Cao"),
        ("F5", "Thông báo", "Sau khi đặt lịch thành công, khách có nhận được thông báo tự động không? Nếu có, kênh nào được ưu tiên: email, SMS, Zalo OA?", "Xác định phạm vi thông báo cho khách.", "Trung bình"),
    ]

    row = 6
    group_rows = []
    for q in questions:
        stt, group, question, reason, priority = q
        is_group = (question == "")

        if is_group:
            ws.merge_cells(f"A{row}:G{row}")
            cell = ws[f"A{row}"]
            cell.value = f"  {stt}. {group}"
            cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
            cell.fill = header_fill(GREEN_LIGHT)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border()
            ws.row_dimensions[row].height = 22
            group_rows.append(row)
        else:
            data = [stt, group, question, reason, priority, "", ""]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = Font(size=9, name="Calibri", color="212529")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border()
                # Alternate row fill
                if row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F2F8F4")
                else:
                    cell.fill = PatternFill("solid", fgColor=WHITE)

            # Priority coloring
            p_cell = ws.cell(row=row, column=5)
            if priority == "Cao":
                p_cell.fill = PatternFill("solid", fgColor="FFE0E0")
                p_cell.font = Font(bold=True, size=9, color=RED_SOFT, name="Calibri")
            elif priority == "Trung bình":
                p_cell.fill = PatternFill("solid", fgColor=GOLD_LIGHT)
                p_cell.font = Font(bold=True, size=9, color=GOLD, name="Calibri")
            else:
                p_cell.fill = PatternFill("solid", fgColor="E8F4FD")
                p_cell.font = Font(bold=True, size=9, color=BLUE_SOFT, name="Calibri")
            p_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

            ws.row_dimensions[row].height = 52

        row += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "* Đề nghị Quý khách vui lòng điền phản hồi vào cột 'Phản hồi của khách' và gửi lại để hai bên tiến hành thống nhất yêu cầu trước khi triển khai."
    ws[f"A{row}"].font = Font(italic=True, size=9, color=GRAY_DARK, name="Calibri")
    ws[f"A{row}"].fill = header_fill(GREEN_PALE)
    ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 28

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A6"

    path = os.path.join(OUTPUT_DIR, "GolfBooking_Questions_Clarification.xlsx")
    wb.save(path)
    print(f"✅ Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
#  FILE 2: QUOTATION & TIMELINE
# ══════════════════════════════════════════════════════════════════════════════
def create_quotation_excel():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Tổng quan báo giá ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Báo giá tổng thể"
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "GOLF TEE TIME BOOKING SYSTEM"
    ws1["A1"].font = Font(bold=True, size=18, color=WHITE, name="Calibri")
    ws1["A1"].fill = header_fill(GREEN_DARK)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 42

    ws1.merge_cells("A2:H2")
    ws1["A2"] = "ĐỀ XUẤT GIÁ & KẾ HOẠCH TRIỂN KHAI — SOFTWARE DEVELOPMENT PROPOSAL"
    ws1["A2"].font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws1["A2"].fill = header_fill(GREEN_MID)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 26

    # Info block
    info = [
        ("Khách hàng", "[Tên sân golf]", "Phiên bản", "1.0"),
        ("Đơn vị thực hiện", "[Tên công ty]", "Ngày lập", "15/04/2026"),
        ("Thời gian thực hiện", "8 tuần (~ 2 tháng)", "Người phụ trách", "[Tên PM]"),
    ]
    for i, (k1, v1, k2, v2) in enumerate(info, 4):
        ws1.merge_cells(f"A{i}:B{i}")
        ws1.merge_cells(f"C{i}:D{i}")
        ws1.merge_cells(f"E{i}:F{i}")
        ws1.merge_cells(f"G{i}:H{i}")

        for col, val, bold, fill_color in [("A", k1, True, "E9ECEF"), ("C", v1, False, WHITE),
                                           ("E", k2, True, "E9ECEF"), ("G", v2, False, WHITE)]:
            c = ws1[f"{col}{i}"]
            c.value = val
            c.font = Font(bold=bold, size=10, name="Calibri", color="212529")
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            c.border = thin_border()
        ws1.row_dimensions[i].height = 20

    # Spacer
    ws1.row_dimensions[7].height = 10

    # ── MODULE TABLE ──────────────────────────────────────────────────────────
    ws1.merge_cells("A8:H8")
    ws1["A8"] = "  CHI TIẾT CÁC MODULE"
    ws1["A8"].font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws1["A8"].fill = header_fill(GREEN_MID)
    ws1["A8"].alignment = Alignment(vertical="center")
    ws1.row_dimensions[8].height = 24

    mod_headers = ["STT", "Module", "Mô tả tóm tắt", "Số ngày công\n(Man-days)", "Đơn giá\n(VND/ngày)", "Thành tiền\n(VND)", "Ghi chú", "Ưu tiên"]
    mod_widths  = [5, 30, 48, 15, 18, 20, 25, 12]
    for i, (h, w) in enumerate(zip(mod_headers, mod_widths), 1):
        c = ws1.cell(row=9, column=i, value=h)
        c.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
        c.fill = header_fill(GREEN_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        set_col_width(ws1, get_column_letter(i), w)
    ws1.row_dimensions[9].height = 36

    modules = [
        # STT, Module, Description, MD, Notes, Priority
        ("1", "iFrame Booking Widget",
         "Giao diện nhúng vào website:\n• Slider ảnh sân golf (trái)\n• Form chọn ngày/giờ/số người (phải)\n• Hiển thị slot khả dụng real-time\n• Responsive (mobile/desktop)",
         10, "Phase 1", "Cao"),
        ("2", "Trang Thanh Toán (Payment Page)",
         "Trang thanh toán độc lập:\n• Tóm tắt booking\n• Form thu thông tin liên hệ (tên, email, SĐT)\n• Tích hợp Airwallex Embedded Elements\n• Hỗ trợ: Visa/MC/Amex/JCB/Discover/Diners\n• Trang xác nhận sau thanh toán",
         12, "Phase 1", "Cao"),
        ("3", "Backend API & Database",
         "API xử lý nghiệp vụ:\n• Quản lý booking (CRUD)\n• Quản lý tee slot & cấu hình\n• Tích hợp Airwallex API\n• Xử lý webhook thanh toán\n• Database schema & migration",
         15, "Phase 1", "Cao"),
        ("4", "Admin — Tee Sheet",
         "Bảng quản lý lịch trực quan:\n• Hiển thị song song 1st Tee & 10th Tee\n• Chuyển ngày nhanh, lọc sáng/chiều\n• Màu sắc theo trạng thái booking\n• Cập nhật real-time (WebSocket)",
         12, "Phase 1", "Cao"),
        ("5", "Admin — Quản lý Booking",
         "Đầy đủ thao tác nghiệp vụ:\n• New / Edit / Move / Copy / Cancel\n• Group Booking\n• Block / Unblock tee time\n• Search & filter booking\n• Waiting list\n• Confirm Letter",
         12, "Phase 1", "Cao"),
        ("6", "Admin — Phân quyền & Đăng nhập",
         "Hệ thống xác thực nội bộ:\n• Đăng nhập email/mật khẩu\n• 3 role: Admin / Sale / Support\n• Admin tạo & quản lý tài khoản\n• Phân quyền theo chức năng",
         5, "Phase 1", "Cao"),
        ("7", "Real-time Notification",
         "Thông báo tức thời:\n• WebSocket: push events đến Admin Panel\n• Webhook: nhận sự kiện từ Airwallex\n• Toast notification & badge trong Admin\n• Cập nhật tee sheet không cần reload",
         6, "Phase 1", "Cao"),
        ("8", "Tính năng nâng cao (TBD)",
         "Phụ thuộc kết quả làm rõ yêu cầu:\n• Squeeze / Unsqueeze\n• Allocate Info\n• Extra Items management\n• Time Sheet / Reports\n• Audit Log",
         8, "Phase 2", "Trung bình"),
        ("9", "Kiểm thử & QA",
         "• Unit test các luồng chính\n• Integration test payment\n• UAT với nhân viên sân golf\n• Bug fixing",
         8, "Toàn dự án", "Cao"),
        ("10", "Triển khai & Go-live",
         "• Cấu hình server / hosting\n• Setup domain, SSL\n• Deploy production\n• Bàn giao tài liệu & hướng dẫn sử dụng",
         4, "Cuối dự án", "Cao"),
    ]

    DAILY_RATE_PLACEHOLDER = "[Đơn giá]"

    row = 10
    total_md = 0
    for i, (stt, name, desc, md, notes, priority) in enumerate(modules):
        total_md += md
        data = [stt, name, desc, md, DAILY_RATE_PLACEHOLDER, "=D{0}*E{0}".format(row) if False else "[Thành tiền]", notes, priority]

        for col_idx, val in enumerate(data, 1):
            c = ws1.cell(row=row, column=col_idx, value=val)
            c.font = Font(size=9, name="Calibri", color="212529")
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if col_idx in [1, 4, 5, 6, 8] else "left")
            c.border = thin_border()
            if row % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F2F8F4")
            else:
                c.fill = PatternFill("solid", fgColor=WHITE)

        # Priority color
        p_cell = ws1.cell(row=row, column=8)
        if priority == "Cao":
            p_cell.fill = PatternFill("solid", fgColor="FFE0E0")
            p_cell.font = Font(bold=True, size=9, color=RED_SOFT, name="Calibri")
        else:
            p_cell.fill = PatternFill("solid", fgColor=GOLD_LIGHT)
            p_cell.font = Font(bold=True, size=9, color=GOLD, name="Calibri")
        p_cell.alignment = Alignment(horizontal="center", vertical="top")

        ws1.row_dimensions[row].height = 70
        row += 1

    # Total row
    for col_idx in range(1, 9):
        c = ws1.cell(row=row, column=col_idx)
        c.border = thin_border()
        c.fill = header_fill(GREEN_PALE)
        c.font = Font(bold=True, size=10, name="Calibri", color=GREEN_DARK)
    ws1.cell(row=row, column=2).value = "TỔNG CỘNG"
    ws1.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=row, column=4).value = total_md
    ws1.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=row, column=5).value = DAILY_RATE_PLACEHOLDER
    ws1.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=row, column=6).value = "[Tổng thành tiền]"
    ws1.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[row].height = 28

    # Notes
    row += 2
    notes_data = [
        "GHI CHÚ & ĐIỀU KHOẢN",
        "1. Thời gian thực hiện: 8 tuần kể từ ngày ký kết hợp đồng và nhận đủ thông tin yêu cầu từ Quý khách.",
        "2. Báo giá chưa bao gồm: phí dịch vụ Airwallex (phụ thuộc volume giao dịch), chi phí server/hosting hàng tháng.",
        "3. Phạm vi tính năng Phase 2 sẽ được xác định sau khi hoàn tất làm rõ yêu cầu ('Danh sách câu hỏi làm rõ').",
        "4. Báo giá có giá trị trong vòng 30 ngày kể từ ngày lập.",
        "5. Thanh toán đề xuất: 40% khi ký hợp đồng — 40% khi nghiệm thu UAT — 20% khi go-live.",
    ]
    ws1.merge_cells(f"A{row}:H{row}")
    ws1[f"A{row}"] = notes_data[0]
    ws1[f"A{row}"].font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    ws1[f"A{row}"].fill = header_fill(GREEN_MID)
    ws1[f"A{row}"].alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[row].height = 22
    row += 1

    for note in notes_data[1:]:
        ws1.merge_cells(f"A{row}:H{row}")
        c = ws1[f"A{row}"]
        c.value = note
        c.font = Font(size=9, name="Calibri", color="212529")
        c.fill = PatternFill("solid", fgColor=GREEN_PALE)
        c.alignment = Alignment(vertical="center", indent=2, wrap_text=True)
        c.border = thin_border()
        ws1.row_dimensions[row].height = 20
        row += 1

    ws1.freeze_panes = "A10"

    # ── Sheet 2: Timeline ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Kế hoạch triển khai")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:L1")
    ws2["A1"] = "KẾ HOẠCH TRIỂN KHAI — PROJECT TIMELINE (8 TUẦN)"
    ws2["A1"].font = Font(bold=True, size=15, color=WHITE, name="Calibri")
    ws2["A1"].fill = header_fill(GREEN_DARK)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 38

    # Week headers
    week_labels = ["Tuần 1", "Tuần 2", "Tuần 3", "Tuần 4", "Tuần 5", "Tuần 6", "Tuần 7", "Tuần 8"]
    week_dates  = ["W1\n(07/04)", "W2\n(14/04)", "W3\n(21/04)", "W4\n(28/04)",
                   "W5\n(05/05)", "W6\n(12/05)", "W7\n(19/05)", "W8\n(26/05)"]

    headers2 = ["STT", "Hạng mục", "Người phụ trách"] + week_labels
    widths2   = [5, 38, 16] + [12]*8

    for i, (h, w) in enumerate(zip(headers2, widths2), 1):
        c = ws2.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
        c.fill = header_fill(GREEN_MID)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        set_col_width(ws2, get_column_letter(i), w)
    ws2.row_dimensions[2].height = 28

    # sub-header: week dates
    for i, d in enumerate(week_dates, 4):
        c = ws2.cell(row=3, column=i, value=d)
        c.font = Font(size=8, italic=True, color=GRAY_DARK, name="Calibri")
        c.fill = PatternFill("solid", fgColor="E9ECEF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws2.row_dimensions[3].height = 22

    GANTT_FILL   = PatternFill("solid", fgColor=GREEN_LIGHT)
    GANTT_EMPTY  = PatternFill("solid", fgColor=WHITE)
    GANTT_PHASE2 = PatternFill("solid", fgColor="A8DADC")
    GANTT_QA     = PatternFill("solid", fgColor=GOLD_LIGHT)
    GANTT_DEPLOY = PatternFill("solid", fgColor="FFE0E0")

    # Data: (stt, name, owner, w1..w8 fill_type: G=gantt, Q=qa, D=deploy, 2=phase2, _=empty)
    timeline = [
        ("─", "PHASE 1 — CORE DEVELOPMENT", "", "","","","","","","",""),
        ("1", "Kickoff & làm rõ yêu cầu", "PM / BA",          "G","G","_","_","_","_","_","_"),
        ("2", "Thiết kế UI/UX (wireframe)", "Designer",        "G","G","G","_","_","_","_","_"),
        ("3", "Setup backend API & Database", "Backend Dev",   "_","G","G","G","_","_","_","_"),
        ("4", "iFrame Booking Widget", "Frontend Dev",         "_","_","G","G","G","_","_","_"),
        ("5", "Trang thanh toán + Airwallex", "Backend + FE",  "_","_","_","G","G","G","_","_"),
        ("6", "Admin — Tee Sheet", "Frontend Dev",             "_","_","_","_","G","G","_","_"),
        ("7", "Admin — Booking Management", "Frontend Dev",    "_","_","_","_","_","G","G","_"),
        ("8", "Admin — Auth & Permission", "Backend Dev",      "_","_","G","G","_","_","_","_"),
        ("9", "Real-time (WebSocket + Webhook)", "Backend Dev","_","_","_","G","G","_","_","_"),
        ("─", "PHASE 2 — ADVANCED FEATURES", "", "","","","","","","",""),
        ("10","Tính năng nâng cao (TBD)", "Dev Team",          "_","_","_","_","_","2","2","_"),
        ("─", "QA & DEPLOY", "", "","","","","","","",""),
        ("11","Kiểm thử & QA", "QA",                          "_","_","_","_","G","G","Q","_"),
        ("12","UAT với Quý khách", "PM + Client",              "_","_","_","_","_","_","Q","Q"),
        ("13","Triển khai & Go-live", "DevOps + PM",           "_","_","_","_","_","_","_","D"),
        ("14","Bàn giao & hướng dẫn", "PM",                    "_","_","_","_","_","_","_","D"),
    ]

    row = 4
    for item in timeline:
        stt  = item[0]
        name = item[1]
        owner = item[2]
        weeks = item[3:]
        is_group = (stt == "─")

        if is_group:
            ws2.merge_cells(f"A{row}:L{row}")
            c = ws2[f"A{row}"]
            c.value = f"  {name}"
            c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
            c.fill = header_fill(GREEN_LIGHT)
            c.alignment = Alignment(vertical="center")
            c.border = thin_border()
            ws2.row_dimensions[row].height = 20
        else:
            for col_idx, val in enumerate([stt, name, owner], 1):
                c = ws2.cell(row=row, column=col_idx, value=val)
                c.font = Font(size=9, name="Calibri", color="212529")
                c.alignment = Alignment(wrap_text=True, vertical="center",
                                        horizontal="center" if col_idx == 1 else "left")
                c.border = thin_border()
                if row % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="F2F8F4")

            for w_idx, w_val in enumerate(weeks, 4):
                c = ws2.cell(row=row, column=w_idx)
                c.border = thin_border()
                if w_val == "G":
                    c.fill = GANTT_FILL
                    c.value = "████"
                    c.font = Font(color=GREEN_LIGHT, size=9, name="Calibri")
                elif w_val == "2":
                    c.fill = GANTT_PHASE2
                    c.value = "████"
                    c.font = Font(color="A8DADC", size=9, name="Calibri")
                elif w_val == "Q":
                    c.fill = GANTT_QA
                    c.value = "████"
                    c.font = Font(color=GOLD_LIGHT, size=9, name="Calibri")
                elif w_val == "D":
                    c.fill = GANTT_DEPLOY
                    c.value = "████"
                    c.font = Font(color="FFE0E0", size=9, name="Calibri")
                else:
                    c.fill = PatternFill("solid", fgColor="F8F9FA") if row % 2 == 0 else PatternFill("solid", fgColor=WHITE)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[row].height = 22

        row += 1

    # Legend
    row += 1
    ws2.merge_cells(f"A{row}:L{row}")
    ws2[f"A{row}"] = "  CHÚ THÍCH MÀU SẮC"
    ws2[f"A{row}"].font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    ws2[f"A{row}"].fill = header_fill(GREEN_MID)
    ws2[f"A{row}"].alignment = Alignment(vertical="center")
    ws2.row_dimensions[row].height = 22
    row += 1

    legends = [
        (GREEN_LIGHT, "Phase 1 — Phát triển tính năng cốt lõi"),
        ("A8DADC", "Phase 2 — Tính năng nâng cao (TBD)"),
        (GOLD_LIGHT, "Kiểm thử & UAT"),
        ("FFE0E0", "Triển khai & Bàn giao"),
    ]
    for fill_c, label in legends:
        ws2.merge_cells(f"A{row}:B{row}")
        c1 = ws2[f"A{row}"]
        c1.fill = PatternFill("solid", fgColor=fill_c)
        c1.value = "████"
        c1.font = Font(color=fill_c, size=10)
        c1.border = thin_border()
        c1.alignment = Alignment(horizontal="center")

        ws2.merge_cells(f"C{row}:L{row}")
        c2 = ws2[f"C{row}"]
        c2.value = label
        c2.font = Font(size=9, name="Calibri")
        c2.fill = PatternFill("solid", fgColor="FAFAFA")
        c2.border = thin_border()
        c2.alignment = Alignment(vertical="center", indent=1)
        ws2.row_dimensions[row].height = 18
        row += 1

    ws2.freeze_panes = "D4"

    path = os.path.join(OUTPUT_DIR, "GolfBooking_Quotation_Timeline.xlsx")
    wb.save(path)
    print(f"✅ Saved: {path}")


if __name__ == "__main__":
    create_questions_excel()
    create_quotation_excel()
    print("\n✅ Done! Both Excel files created.")
